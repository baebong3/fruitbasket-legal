"""KAMIS 농산물 가격 데이터 엑셀 누적 저장 모듈

매일 수집한 데이터를 하나의 엑셀 파일에 누적 저장합니다.
- 월별 시트로 자동 분리 (예: 2025-01, 2025-02)
- 중복 데이터 자동 방지 (날짜 + 품목코드 + 품종코드 + 등급코드 기준)
- 요약 시트 자동 업데이트
"""

import logging
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)

KST = timezone(timedelta(hours=9))

# 엑셀 컬럼 정의
COLUMNS = [
    ("date", "날짜"),
    ("category_name", "부류"),
    ("item_code", "품목코드"),
    ("item_name", "품목명"),
    ("kind_code", "품종코드"),
    ("kind_name", "품종명"),
    ("rank", "등급"),
    ("rank_code", "등급코드"),
    ("unit", "단위"),
    ("price", "가격(원)"),
    ("market_name", "지역"),
]

# 스타일 정의
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F6B3B", end_color="1F6B3B", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _get_output_filepath(output_dir: Path) -> Path:
    """누적 저장용 엑셀 파일 경로 반환"""
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / "kamis_prices.xlsx"


def _get_month_sheet_name(date_str: str) -> str:
    """날짜 문자열에서 월별 시트 이름 생성 (예: '2025-01-15' -> '2025-01')"""
    try:
        parts = date_str.split("-")
        return f"{parts[0]}-{parts[1]}"
    except (IndexError, ValueError):
        return "기타"


def _apply_header_style(ws, col_count: int):
    """헤더 행에 스타일 적용"""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _write_header(ws):
    """시트에 헤더 행 작성"""
    for col_idx, (_, col_name) in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    _apply_header_style(ws, len(COLUMNS))
    ws.freeze_panes = "A2"


def _get_existing_keys(ws) -> set:
    """시트에서 기존 데이터의 고유 키 세트 반환 (중복 방지용)

    키: (날짜, 품목코드, 품종코드, 등급코드, 지역)
    """
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        # date, _, item_code, _, kind_code, _, _, rank_code, _, _, market_name
        key = (str(row[0]), str(row[2]), str(row[4]), str(row[7]), str(row[10]))
        keys.add(key)
    return keys


def _make_key(record: dict) -> tuple:
    """레코드에서 고유 키 생성"""
    return (
        str(record.get("date", "")),
        str(record.get("item_code", "")),
        str(record.get("kind_code", "")),
        str(record.get("rank_code", "")),
        str(record.get("market_name", "")),
    )


def _auto_fit_columns(ws):
    """컬럼 너비 자동 조정"""
    for col_idx, (_, col_name) in enumerate(COLUMNS, 1):
        max_len = len(col_name)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)), 40))
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = max_len + 4


def _update_summary_sheet(wb: Workbook, all_data: dict[str, list[dict]]):
    """요약 시트 생성/업데이트

    Args:
        wb: 워크북 객체
        all_data: {시트이름: [레코드, ...]} 형태의 데이터
    """
    summary_name = "요약"
    if summary_name in wb.sheetnames:
        del wb[summary_name]

    ws = wb.create_sheet(title=summary_name, index=0)

    # 헤더
    headers = ["월", "총 데이터 건수", "품목 수", "최종 업데이트"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    now_str = datetime.now(KST).strftime("%Y-%m-%d %H:%M")

    row_idx = 2
    total_records = 0
    for sheet_name in sorted(wb.sheetnames):
        if sheet_name == summary_name:
            continue
        sheet = wb[sheet_name]
        record_count = sheet.max_row - 1 if sheet.max_row > 1 else 0
        total_records += record_count

        # 해당 월의 고유 품목 수 계산
        item_names = set()
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
            if row[0]:
                item_names.add(row[0])

        ws.cell(row=row_idx, column=1, value=sheet_name)
        ws.cell(row=row_idx, column=2, value=record_count)
        ws.cell(row=row_idx, column=3, value=len(item_names))
        ws.cell(row=row_idx, column=4, value=now_str)
        row_idx += 1

    # 합계 행
    ws.cell(row=row_idx, column=1, value="합계").font = Font(bold=True)
    ws.cell(row=row_idx, column=2, value=total_records).font = Font(bold=True)

    # 컬럼 너비
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 22
    ws.freeze_panes = "A2"


def save_to_excel(price_data: list[dict], output_dir: Path | None = None) -> Path:
    """가격 데이터를 엑셀 파일에 누적 저장

    Args:
        price_data: kamis_collector에서 수집한 정규화된 가격 데이터 리스트
        output_dir: 출력 디렉토리 (기본값: ./output)

    Returns:
        저장된 엑셀 파일 경로
    """
    if output_dir is None:
        output_dir = Path(__file__).parent / "output"

    filepath = _get_output_filepath(output_dir)

    # 기존 파일이 있으면 로드, 없으면 새로 생성
    if filepath.exists():
        wb = load_workbook(filepath)
        logger.info("기존 엑셀 파일 로드: %s", filepath)
    else:
        wb = Workbook()
        wb.remove(wb.active)
        logger.info("새 엑셀 파일 생성: %s", filepath)

    # 월별로 데이터 분류
    monthly_data: dict[str, list[dict]] = {}
    for record in price_data:
        sheet_name = _get_month_sheet_name(record.get("date", ""))
        monthly_data.setdefault(sheet_name, []).append(record)

    total_new = 0
    total_dup = 0

    for sheet_name, records in monthly_data.items():
        # 시트 가져오기 또는 새로 생성
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            _write_header(ws)
            logger.info("새 시트 생성: %s", sheet_name)

        # 기존 키 세트 로드 (중복 방지)
        existing_keys = _get_existing_keys(ws)
        next_row = ws.max_row + 1 if ws.max_row > 1 else 2

        new_count = 0
        dup_count = 0

        for record in records:
            key = _make_key(record)
            if key in existing_keys:
                dup_count += 1
                continue

            existing_keys.add(key)
            for col_idx, (field, _) in enumerate(COLUMNS, 1):
                value = record.get(field, "")
                if field == "price" and value is not None:
                    try:
                        value = int(value)
                    except (ValueError, TypeError):
                        pass
                ws.cell(row=next_row, column=col_idx, value=value)
            next_row += 1
            new_count += 1

        total_new += new_count
        total_dup += dup_count

        if new_count > 0:
            _auto_fit_columns(ws)
            logger.info("[%s] 신규 %d건 추가 (중복 %d건 건너뜀)", sheet_name, new_count, dup_count)

    # 요약 시트 업데이트
    _update_summary_sheet(wb, monthly_data)

    wb.save(filepath)
    logger.info("엑셀 저장 완료: %s (신규 %d건, 중복 %d건)", filepath, total_new, total_dup)
    return filepath


if __name__ == "__main__":
    # 테스트용 샘플 데이터
    sample_data = [
        {
            "date": "2025-01-15",
            "category_code": "400",
            "category_name": "과일류",
            "item_code": "411",
            "item_name": "사과",
            "kind_code": "01",
            "kind_name": "후지",
            "rank": "상품",
            "rank_code": "04",
            "unit": "10 개",
            "price": 25000,
            "market_name": "서울",
        },
        {
            "date": "2025-01-15",
            "category_code": "400",
            "category_name": "과일류",
            "item_code": "412",
            "item_name": "배",
            "kind_code": "01",
            "kind_name": "신고",
            "rank": "상품",
            "rank_code": "04",
            "unit": "10 개",
            "price": 35000,
            "market_name": "서울",
        },
    ]
    filepath = save_to_excel(sample_data)
    print(f"테스트 저장 완료: {filepath}")
