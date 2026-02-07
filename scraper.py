"""나라장터 입찰공고정보서비스 - 용역 입찰공고 수집기"""

import json
import logging
import os
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path
import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

BASE_URL = "http://apis.data.go.kr/1230000/BidPublicInfoService/getBidPblancListInfoServc"
KST = timezone(timedelta(hours=9))

COLUMNS = [
    ("bidNtceNo", "공고번호"),
    ("bidNtceNm", "공고명"),
    ("ntceInsttNm", "발주기관"),
    ("bidNtceOrd", "공고차수"),
    ("bidClseDt", "입찰마감일시"),
    ("presmptPrce", "추정가격"),
    ("asignBdgtAmt", "배정예산액"),
    ("bidNtceDt", "공고일시"),
    ("rbidPermsnYn", "재입찰허용여부"),
    ("bidNtceDtlUrl", "공고URL"),
]


def load_config() -> dict:
    config_path = Path(__file__).parent / "config.json"
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def get_api_key() -> str:
    key = os.environ.get("G2B_API_KEY")
    if not key:
        logger.error("G2B_API_KEY 환경변수가 설정되지 않았습니다.")
        sys.exit(1)
    return key


def build_params(config: dict, target_date: datetime) -> dict:
    start_hour = config.get("search_period", {}).get("start_hour", "0000")
    end_hour = config.get("search_period", {}).get("end_hour", "2359")
    date_str = target_date.strftime("%Y%m%d")

    return {
        "pageNo": 1,
        "numOfRows": config.get("num_of_rows", 999),
        "inqryDiv": config.get("inqry_div", "1"),
        "inqryBgnDt": f"{date_str}{start_hour}",
        "inqryEndDt": f"{date_str}{end_hour}",
        "type": "json",
    }


def fetch_bids(api_key: str, config: dict, target_date: datetime) -> list[dict]:
    params = build_params(config, target_date)
    url = f"{BASE_URL}?ServiceKey={api_key}"
    all_items = []
    page = 1

    while True:
        params["pageNo"] = page
        logger.info("API 호출 중... (페이지 %d)", page)

        try:
            resp = requests.get(url, params=params, timeout=30)
            resp.raise_for_status()
        except requests.RequestException as e:
            logger.error("API 요청 실패: %s", e)
            break

        try:
            data = resp.json()
        except ValueError:
            logger.error("JSON 파싱 실패. 응답: %s", resp.text[:500])
            break

        response = data.get("response", {})
        header = response.get("header", {})

        if header.get("resultCode") != "00":
            logger.error("API 오류: %s", header.get("resultMsg", "알 수 없는 오류"))
            break

        body = response.get("body", {})
        items = body.get("items", [])

        if not items:
            if page == 1:
                logger.info("조회 결과가 없습니다.")
            break

        all_items.extend(items)
        total_count = body.get("totalCount", 0)
        logger.info("페이지 %d 조회 완료 (%d/%d건)", page, len(all_items), total_count)

        if len(all_items) >= total_count:
            break

        page += 1

    logger.info("총 %d건 조회 완료", len(all_items))
    return all_items


def filter_by_keywords(items: list[dict], keywords: list[str]) -> dict[str, list[dict]]:
    results = {}
    for keyword in keywords:
        matched = [
            item for item in items
            if keyword in (item.get("bidNtceNm") or "")
        ]
        results[keyword] = matched
        logger.info("검색어 '%s': %d건 매칭", keyword, len(matched))
    return results


def write_excel(filtered: dict[str, list[dict]], target_date: datetime, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    date_str = target_date.strftime("%Y%m%d")
    filepath = output_dir / f"g2b_용역공고_{date_str}.xlsx"

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for keyword, items in filtered.items():
        ws = wb.create_sheet(title=keyword)

        # Write header
        for col_idx, (_, col_name) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Write data
        for row_idx, item in enumerate(items, 2):
            for col_idx, (field, _) in enumerate(COLUMNS, 1):
                value = item.get(field, "")
                if field == "presmptPrce" or field == "asignBdgtAmt":
                    try:
                        value = int(value) if value else ""
                    except (ValueError, TypeError):
                        pass
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths (approximate)
        for col_idx, (_, col_name) in enumerate(COLUMNS, 1):
            max_len = len(col_name)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 60))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 4

        # Freeze header row
        ws.freeze_panes = "A2"

    # Add summary sheet
    ws_summary = wb.create_sheet(title="요약", index=0)
    ws_summary.cell(row=1, column=1, value="검색어").font = Font(bold=True)
    ws_summary.cell(row=1, column=2, value="건수").font = Font(bold=True)
    ws_summary.cell(row=1, column=3, value="조회일자").font = Font(bold=True)
    for idx, (keyword, items) in enumerate(filtered.items(), 2):
        ws_summary.cell(row=idx, column=1, value=keyword)
        ws_summary.cell(row=idx, column=2, value=len(items))
        ws_summary.cell(row=idx, column=3, value=date_str)
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 10
    ws_summary.column_dimensions["C"].width = 15

    wb.save(filepath)
    logger.info("엑셀 저장 완료: %s", filepath)
    return filepath


def main():
    config = load_config()
    api_key = get_api_key()
    keywords = config.get("keywords", ["조사", "연구"])

    # Use today in KST
    target_date = datetime.now(KST)
    logger.info("조회 대상일: %s", target_date.strftime("%Y-%m-%d"))
    logger.info("검색어: %s", keywords)

    items = fetch_bids(api_key, config, target_date)

    if not items:
        logger.info("조회된 공고가 없어 엑셀을 생성하지 않습니다.")
        return

    filtered = filter_by_keywords(items, keywords)
    total_matched = sum(len(v) for v in filtered.values())

    if total_matched == 0:
        logger.info("검색어에 매칭되는 공고가 없습니다.")

    output_dir = Path(__file__).parent / "output"
    filepath = write_excel(filtered, target_date, output_dir)
    logger.info("완료. 파일: %s", filepath)


if __name__ == "__main__":
    main()
